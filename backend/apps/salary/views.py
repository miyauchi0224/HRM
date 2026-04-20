import csv
import io
from datetime import date
from decimal import Decimal

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from apps.accounts.permissions import IsNotCustomer, IsHR, IsAccounting
from rest_framework.response import Response
from django.http import HttpResponse
from django.utils import timezone

from .models import SalaryGrade, Allowance, EmployeeAllowance, Payslip
from .serializers import (
    SalaryGradeSerializer, AllowanceSerializer,
    EmployeeAllowanceSerializer, PayslipSerializer
)
from apps.employees.models import Employee
from apps.attendance.models import AttendanceRecord


# ── 社会保険・税率定数（概算） ─────────────────────────────────────────────
HEALTH_INSURANCE_RATE     = Decimal('0.0500')   # 健康保険 5.00%（労働者負担）
PENSION_RATE              = Decimal('0.0915')   # 厚生年金 9.15%
EMPLOYMENT_INSURANCE_RATE = Decimal('0.006')    # 雇用保険 0.6%
NURSING_INSURANCE_RATE    = Decimal('0.0091')   # 介護保険 0.91%（40歳以上）
OVERTIME_RATE             = Decimal('1.25')     # 残業割増率


class SalaryGradeViewSet(viewsets.ModelViewSet):
    """給与等級マスタ：参照は社員以上、編集は人事以上"""
    queryset           = SalaryGrade.objects.all()
    serializer_class   = SalaryGradeSerializer
    permission_classes = [IsNotCustomer]

    def get_permissions(self):
        if self.request.method in ('POST', 'PUT', 'PATCH', 'DELETE'):
            return [IsHR()]
        return [IsAuthenticated()]


class AllowanceViewSet(viewsets.ModelViewSet):
    queryset           = Allowance.objects.all()
    serializer_class   = AllowanceSerializer
    permission_classes = [IsNotCustomer]

    def get_permissions(self):
        if self.request.method in ('POST', 'PUT', 'PATCH', 'DELETE'):
            return [IsHR()]
        return [IsAuthenticated()]


class EmployeeAllowanceViewSet(viewsets.ModelViewSet):
    serializer_class   = EmployeeAllowanceSerializer
    permission_classes = [IsNotCustomer]

    def get_queryset(self):
        user = self.request.user
        if user.is_accounting:  # 経理・人事・管理者は全員分を参照可
            emp_id = self.request.query_params.get('employee_id')
            if emp_id:
                return EmployeeAllowance.objects.filter(employee_id=emp_id)
            return EmployeeAllowance.objects.all()
        return EmployeeAllowance.objects.filter(employee__user=user)


class PayslipViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class   = PayslipSerializer
    permission_classes = [IsNotCustomer]

    def get_queryset(self):
        user = self.request.user
        if user.is_accounting:  # 経理・人事・管理者は全員分を参照可
            emp_id = self.request.query_params.get('employee_id')
            qs = Payslip.objects.select_related('employee')
            if emp_id:
                return qs.filter(employee_id=emp_id)
            return qs
        return Payslip.objects.filter(employee__user=user).select_related('employee')

    @action(detail=False, methods=['get'], url_path='template-csv', permission_classes=[IsAccounting])
    def template_csv(self, request):
        """
        GET /api/v1/salary/payslips/template-csv/
        給与明細一括登録用 CSV テンプレートをダウンロード（経理・人事・管理者のみ）
        """
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            '社員番号', '年', '月',
            '基本給', '技術手当', '出向手当', '住宅手当', '残業手当',
            '通勤手当', '家族手当', '資格手当', '役職手当', '特別手当',
            '皆勤手当', '精勤手当', '時間外手当',
            '健康保険料', '厚生年金', '雇用保険料', '介護保険料',
            '財形貯蓄', '社宅寮費', '組合費', '共済会費', '持株会拠出金',
            'その他控除', '所得税', '住民税',
            '出勤日数', '欠勤日数', '有給取得日数',
            '締め日', '支給日', '備考',
        ])
        # サンプル行
        today = date.today()
        writer.writerow([
            'EMP001', today.year, today.month,
            300000, 20000, 0, 15000, 10000,
            12000, 5000, 0, 0, 0,
            0, 0, 0,
            15000, 27450, 1800, 0,
            0, 0, 0, 0, 0,
            0, 8000, 12000,
            20, 0, 0,
            f'{today.year}/{today.month:02d}/15',
            f'{today.year}/{today.month:02d}/25',
            '',
        ])
        response = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="payslip_upload_template.csv"'
        return response

    @action(detail=False, methods=['post'], url_path='import-csv',
            parser_classes=[MultiPartParser, FormParser],
            permission_classes=[IsAccounting])
    def import_csv(self, request):
        """
        POST /api/v1/salary/payslips/import-csv/
        給与明細を CSV で一括登録・更新（経理・人事・管理者のみ）

        同一社員・年・月の明細が既存の場合は上書き更新する。
        合計は自動計算される（総支給額・総控除額・差引支給額）。
        """
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'ファイルが指定されていません'}, status=status.HTTP_400_BAD_REQUEST)

        decoded = file.read().decode('utf-8-sig')
        reader  = csv.DictReader(io.StringIO(decoded))

        created = 0
        updated = 0
        errors  = []

        def to_int(val, default=0):
            v = str(val).strip().replace(',', '')
            return int(v) if v else default

        def parse_date(val):
            if not val or not val.strip():
                return None
            return date.fromisoformat(val.strip().replace('/', '-'))

        for i, row in enumerate(reader, start=2):
            try:
                emp_number = row.get('社員番号', '').strip()
                year  = int(row.get('年', '').strip())
                month = int(row.get('月', '').strip())
                if not emp_number:
                    raise ValueError('社員番号は必須です')

                employee = Employee.objects.filter(employee_number=emp_number).first()
                if not employee:
                    raise ValueError(f'社員番号が見つかりません: {emp_number}')

                payslip, is_new = Payslip.objects.get_or_create(
                    employee=employee, year=year, month=month,
                )
                # 支給項目
                payslip.base_salary                 = to_int(row.get('基本給', 0))
                payslip.technical_allowance         = to_int(row.get('技術手当', 0))
                payslip.secondment_allowance        = to_int(row.get('出向手当', 0))
                payslip.housing_allowance           = to_int(row.get('住宅手当', 0))
                payslip.overtime_pay                = to_int(row.get('残業手当', 0))
                payslip.commute_allowance           = to_int(row.get('通勤手当', 0))
                payslip.family_allowance            = to_int(row.get('家族手当', 0))
                payslip.certification_allowance     = to_int(row.get('資格手当', 0))
                payslip.position_allowance          = to_int(row.get('役職手当', 0))
                payslip.special_allowance           = to_int(row.get('特別手当', 0))
                payslip.perfect_attendance_allowance= to_int(row.get('皆勤手当', 0))
                payslip.diligence_allowance         = to_int(row.get('精勤手当', 0))
                payslip.extra_overtime_pay          = to_int(row.get('時間外手当', 0))
                # 控除項目
                payslip.health_insurance            = to_int(row.get('健康保険料', 0))
                payslip.pension                     = to_int(row.get('厚生年金', 0))
                payslip.employment_insurance        = to_int(row.get('雇用保険料', 0))
                payslip.nursing_insurance           = to_int(row.get('介護保険料', 0))
                payslip.property_savings            = to_int(row.get('財形貯蓄', 0))
                payslip.company_housing_fee         = to_int(row.get('社宅寮費', 0))
                payslip.union_fee                   = to_int(row.get('組合費', 0))
                payslip.mutual_aid_fee              = to_int(row.get('共済会費', 0))
                payslip.employee_stock_contribution = to_int(row.get('持株会拠出金', 0))
                payslip.other_deductions            = to_int(row.get('その他控除', 0))
                payslip.income_tax                  = to_int(row.get('所得税', 0))
                payslip.resident_tax                = to_int(row.get('住民税', 0))
                # 勤怠
                payslip.work_days      = to_int(row.get('出勤日数', 0))
                payslip.absence_days   = to_int(row.get('欠勤日数', 0))
                payslip.paid_leave_days= to_int(row.get('有給取得日数', 0))
                # 管理情報
                payslip.cutoff_date  = parse_date(row.get('締め日'))
                payslip.payment_date = parse_date(row.get('支給日'))
                payslip.note         = row.get('備考', '').strip()

                payslip.recompute_totals()
                payslip.save()

                if is_new:
                    created += 1
                else:
                    updated += 1

            except Exception as e:
                errors.append({'row': i, 'error': str(e)})

        return Response({
            'created': created,
            'updated': updated,
            'errors':  len(errors),
            'error_details': errors,
        })

    @action(detail=False, methods=['post'], url_path='calculate', permission_classes=[IsAccounting])
    def calculate(self, request):
        """
        POST /api/v1/salary/payslips/calculate/
        Body: { "year": 2026, "month": 4, "employee_id": "uuid"（省略で全員） }
        指定月の給与を計算して Payslip を生成（または更新）
        """
        year  = request.data.get('year')
        month = request.data.get('month')
        if not year or not month:
            return Response({'error': 'year と month は必須です'}, status=status.HTTP_400_BAD_REQUEST)

        emp_id    = request.data.get('employee_id')
        employees = Employee.objects.all()
        if emp_id:
            employees = employees.filter(id=emp_id)

        results = []
        for emp in employees:
            payslip = _calculate_payslip(emp, int(year), int(month))
            results.append(PayslipSerializer(payslip).data)

        return Response({'calculated': len(results), 'payslips': results})

    @action(detail=True, methods=['patch'], url_path='update', permission_classes=[IsHR])
    def update_payslip(self, request, pk=None):
        """
        PATCH /api/v1/salary/payslips/{id}/update/
        手当・控除の個別項目を人事担当者が手動修正し、合計を再計算する
        """
        payslip = self.get_object()
        # 編集可能フィールドのみ受け付ける
        editable = [
            'technical_allowance', 'secondment_allowance', 'housing_allowance',
            'overtime_pay', 'commute_allowance', 'family_allowance',
            'certification_allowance', 'position_allowance', 'special_allowance',
            'perfect_attendance_allowance', 'diligence_allowance', 'extra_overtime_pay',
            'health_insurance', 'pension', 'employment_insurance', 'nursing_insurance',
            'property_savings', 'company_housing_fee', 'union_fee', 'mutual_aid_fee',
            'employee_stock_contribution', 'other_deductions',
            'income_tax', 'resident_tax',
            'work_days', 'absence_days', 'paid_leave_days',
            'cutoff_date', 'payment_date', 'note', 'status',
        ]
        for field in editable:
            if field in request.data:
                setattr(payslip, field, request.data[field])

        payslip.recompute_totals()
        payslip.save()
        return Response(PayslipSerializer(payslip).data)

    @action(detail=True, methods=['get'], url_path='download')
    def download(self, request, pk=None):
        """
        GET /api/v1/salary/payslips/{id}/download/
        給与明細を XLSX 形式でダウンロード（給与明細書レイアウト）
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        payslip = self.get_object()
        if not request.user.is_accounting and payslip.employee.user != request.user:
            return Response({'error': '権限がありません'}, status=status.HTTP_403_FORBIDDEN)

        emp = payslip.employee
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = '給与明細書'

        # ────── スタイル定義 ──────
        BLUE   = '2E75B6'
        LBLUE  = 'D6E4F7'
        GREEN  = 'E2EFDA'
        RED    = 'FCE4D6'
        GRAY   = 'F2F2F2'
        BOLD14 = Font(bold=True, size=14)
        BOLD11 = Font(bold=True, size=11)
        BOLD10 = Font(bold=True, size=10)
        NORM9  = Font(size=9)
        HDR_F  = Font(color='FFFFFF', bold=True, size=10)
        HDR_F2 = Font(bold=True, size=10)

        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hdr_fill(color):
            return PatternFill(fill_type='solid', fgColor=color)

        def cell(row, col, value, font=None, fill=None, align='left', border_=True, number_format=None):
            c = ws.cell(row=row, column=col, value=value)
            if font:   c.font = font
            if fill:   c.fill = fill
            if border_: c.border = border
            c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
            if number_format: c.number_format = number_format
            return c

        def money(row, col, value, fill=None):
            cell(row, col, value, font=NORM9, fill=fill, align='right', number_format='#,##0')

        # ────── タイトル ──────
        ws.merge_cells('A1:H1')
        cell(1, 1, '給与明細書', font=BOLD14, fill=hdr_fill(BLUE), align='center')
        ws['A1'].font = Font(color='FFFFFF', bold=True, size=14)

        # ────── ヘッダー情報 ──────
        ws.merge_cells('A2:D2')
        cell(2, 1, f'{payslip.year}年{payslip.month}月分', font=BOLD11, align='center')
        if payslip.payment_date:
            ws.merge_cells('E2:H2')
            cell(2, 5, f'支給日: {payslip.payment_date}', font=NORM9, align='right')

        ROW = 3
        info_rows = [
            ('社員番号', emp.employee_number, '部署',     emp.department),
            ('氏名',     emp.full_name,       '役職',     emp.position),
            ('口座情報', PayslipSerializer().get_bank_info(payslip), '備考', payslip.note),
        ]
        for lbl1, val1, lbl2, val2 in info_rows:
            cell(ROW, 1, lbl1,  font=BOLD10, fill=hdr_fill(LBLUE), align='center')
            ws.merge_cells(f'B{ROW}:D{ROW}')
            cell(ROW, 2, val1,  font=NORM9)
            cell(ROW, 5, lbl2,  font=BOLD10, fill=hdr_fill(LBLUE), align='center')
            ws.merge_cells(f'F{ROW}:H{ROW}')
            cell(ROW, 6, val2,  font=NORM9)
            ROW += 1

        # ────── 勤怠情報 ──────
        ROW += 1
        ws.merge_cells(f'A{ROW}:H{ROW}')
        cell(ROW, 1, '【勤怠情報】', font=BOLD10, fill=hdr_fill(LBLUE))
        ROW += 1
        for lbl, val in [
            ('出勤日数', f'{payslip.work_days}日'),
            ('欠勤日数', f'{payslip.absence_days}日'),
            ('有給取得', f'{payslip.paid_leave_days}日'),
            ('締め日', str(payslip.cutoff_date) if payslip.cutoff_date else '—'),
        ]:
            col_offset = [
                ('出勤日数', 1), ('欠勤日数', 3), ('有給取得', 5), ('締め日', 7),
            ]
        attendance_items = [
            ('出勤日数', f'{payslip.work_days}日', 1),
            ('欠勤日数', f'{payslip.absence_days}日', 3),
            ('有給取得日数', f'{payslip.paid_leave_days}日', 5),
            ('締め日', str(payslip.cutoff_date) if payslip.cutoff_date else '—', 7),
        ]
        for lbl, val, col in attendance_items:
            cell(ROW, col,     lbl, font=BOLD10, fill=hdr_fill(GRAY), align='center')
            cell(ROW, col + 1, val, font=NORM9,  align='center')

        # ────── 支給・控除 ──────
        ROW += 2
        # ヘッダー行
        ws.merge_cells(f'A{ROW}:D{ROW}')
        cell(ROW, 1, '【支給】', font=HDR_F, fill=hdr_fill(BLUE), align='center')
        ws.merge_cells(f'E{ROW}:H{ROW}')
        cell(ROW, 5, '【控除】', font=HDR_F, fill=hdr_fill(BLUE), align='center')
        ROW += 1

        PAYMENTS = [
            ('基本給',                    payslip.base_salary),
            ('技術手当',                  payslip.technical_allowance),
            ('出向手当',                  payslip.secondment_allowance),
            ('住宅手当',                  payslip.housing_allowance),
            ('残業手当',                  payslip.overtime_pay),
            ('通勤手当（交通費）',         payslip.commute_allowance),
            ('家族手当',                  payslip.family_allowance),
            ('資格手当',                  payslip.certification_allowance),
            ('役職手当',                  payslip.position_allowance),
            ('特別手当（賞与・臨時）',     payslip.special_allowance),
            ('皆勤手当',                  payslip.perfect_attendance_allowance),
            ('精勤手当',                  payslip.diligence_allowance),
            ('時間外手当（深夜・休日）',   payslip.extra_overtime_pay),
        ]
        DEDUCTIONS = [
            ('健康保険料',        payslip.health_insurance),
            ('厚生年金',          payslip.pension),
            ('雇用保険料',        payslip.employment_insurance),
            ('介護保険料',        payslip.nursing_insurance),
            ('社会保険料合計',    payslip.social_insurance_total),
            ('財形貯蓄',          payslip.property_savings),
            ('社宅・寮費',        payslip.company_housing_fee),
            ('組合費',            payslip.union_fee),
            ('共済会費',          payslip.mutual_aid_fee),
            ('持株会拠出金',      payslip.employee_stock_contribution),
            ('その他控除',        payslip.other_deductions),
            ('所得税',            payslip.income_tax),
            ('住民税',            payslip.resident_tax),
        ]

        max_rows = max(len(PAYMENTS), len(DEDUCTIONS))
        p_fill = hdr_fill(GRAY)
        d_fill = hdr_fill(GRAY)

        for i in range(max_rows):
            r = ROW + i
            # 支給側
            if i < len(PAYMENTS):
                lbl, val = PAYMENTS[i]
                cell(r, 1, lbl, font=NORM9, fill=p_fill, align='left')
                ws.merge_cells(f'B{r}:C{r}')
                money(r, 2, val)
                cell(r, 4, '', fill=p_fill)
            # 控除側
            if i < len(DEDUCTIONS):
                lbl, val = DEDUCTIONS[i]
                bg = hdr_fill(LBLUE) if lbl == '社会保険料合計' else d_fill
                cell(r, 5, lbl, font=NORM9 if lbl != '社会保険料合計' else BOLD10,
                     fill=bg, align='left')
                ws.merge_cells(f'F{r}:G{r}')
                money(r, 6, val, fill=bg if lbl == '社会保険料合計' else None)
                cell(r, 8, '', fill=d_fill)

        # 合計行
        ROW += max_rows
        ws.merge_cells(f'A{ROW}:C{ROW}')
        cell(ROW, 1, '支給合計', font=HDR_F2, fill=hdr_fill(GREEN), align='left')
        money(ROW, 4, payslip.gross_salary, fill=hdr_fill(GREEN))

        ws.merge_cells(f'E{ROW}:G{ROW}')
        cell(ROW, 5, '控除合計', font=HDR_F2, fill=hdr_fill(RED), align='left')
        money(ROW, 8, payslip.total_deductions, fill=hdr_fill(RED))

        # 差引支給額
        ROW += 2
        ws.merge_cells(f'A{ROW}:F{ROW}')
        cell(ROW, 1, '差引支給額（手取）', font=Font(bold=True, size=12), fill=hdr_fill(BLUE), align='center')
        ws['A' + str(ROW)].font = Font(color='FFFFFF', bold=True, size=12)
        ws.merge_cells(f'G{ROW}:H{ROW}')
        money(ROW, 7, payslip.net_salary, fill=hdr_fill(BLUE))
        ws.cell(row=ROW, column=7).font = Font(color='FFFFFF', bold=True, size=12)

        # 列幅調整
        col_widths = [16, 12, 10, 12, 16, 12, 10, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        # 行高さ
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'payslip_{payslip.year}{payslip.month:02d}_{emp.employee_number}.xlsx'
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['get'], url_path='download-pdf')
    def download_pdf(self, request, pk=None):
        """
        GET /api/v1/salary/payslips/{id}/download-pdf/
        給与明細を PDF 形式でダウンロード
        """
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from config.pdf_fonts import register_fonts
        register_fonts()  # 冪等（起動時登録済みの場合は即return）

        payslip = self.get_object()
        if not request.user.is_accounting and payslip.employee.user != request.user:
            return Response({'error': '権限がありません'}, status=status.HTTP_403_FORBIDDEN)

        emp = payslip.employee
        FONT = 'HeiseiKakuGo-W5'

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=15*mm, bottomMargin=15*mm)

        title_s  = ParagraphStyle('t', fontName=FONT, fontSize=14, leading=20, spaceAfter=6, alignment=1)
        sub_s    = ParagraphStyle('s', fontName=FONT, fontSize=9,  leading=13)
        label_s  = ParagraphStyle('l', fontName=FONT, fontSize=8,  leading=12)

        fmt = lambda n: f'¥{int(n):,}'

        elements = [
            Paragraph('給与明細書', title_s),
            Paragraph(f'{payslip.year}年{payslip.month}月分　支給日: {payslip.payment_date or "—"}', sub_s),
            Spacer(1, 4*mm),
        ]

        # 社員情報テーブル
        info_data = [
            ['社員番号', emp.employee_number, '部署', emp.department],
            ['氏名',     emp.full_name,       '役職', emp.position],
            ['口座情報', PayslipSerializer().get_bank_info(payslip), '備考', payslip.note or '—'],
        ]
        info_tbl = Table(info_data, colWidths=[25*mm, 55*mm, 20*mm, 55*mm])
        info_tbl.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), FONT),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D6E4F7')),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#D6E4F7')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWHEIGHT', (0, 0), (-1, -1), 14),
        ]))
        elements += [info_tbl, Spacer(1, 4*mm)]

        # 勤怠テーブル
        att_data = [
            ['出勤日数', f'{payslip.work_days}日', '欠勤日数', f'{payslip.absence_days}日',
             '有給取得', f'{payslip.paid_leave_days}日', '締め日', str(payslip.cutoff_date or '—')],
        ]
        att_tbl = Table(att_data, colWidths=[22*mm, 18*mm]*4)
        att_tbl.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), FONT),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#F2F2F2')),
            ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#F2F2F2')),
            ('BACKGROUND', (4, 0), (4, 0), colors.HexColor('#F2F2F2')),
            ('BACKGROUND', (6, 0), (6, 0), colors.HexColor('#F2F2F2')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        elements += [att_tbl, Spacer(1, 4*mm)]

        # 支給・控除テーブル
        BLUE_H = colors.HexColor('#2E75B6')
        LBLUE  = colors.HexColor('#D6E4F7')
        GRAY   = colors.HexColor('#F2F2F2')
        GREEN  = colors.HexColor('#E2EFDA')
        RED_   = colors.HexColor('#FCE4D6')

        PAYMENTS = [
            ('基本給',                   payslip.base_salary),
            ('技術手当',                 payslip.technical_allowance),
            ('出向手当',                 payslip.secondment_allowance),
            ('住宅手当',                 payslip.housing_allowance),
            ('残業手当',                 payslip.overtime_pay),
            ('通勤手当（交通費）',        payslip.commute_allowance),
            ('家族手当',                 payslip.family_allowance),
            ('資格手当',                 payslip.certification_allowance),
            ('役職手当',                 payslip.position_allowance),
            ('特別手当（賞与・臨時）',    payslip.special_allowance),
            ('皆勤手当',                 payslip.perfect_attendance_allowance),
            ('精勤手当',                 payslip.diligence_allowance),
            ('時間外手当（深夜・休日）',  payslip.extra_overtime_pay),
        ]
        DEDUCTIONS = [
            ('健康保険料',       payslip.health_insurance),
            ('厚生年金',         payslip.pension),
            ('雇用保険料',       payslip.employment_insurance),
            ('介護保険料',       payslip.nursing_insurance),
            ('社会保険料合計',   payslip.social_insurance_total),
            ('財形貯蓄',         payslip.property_savings),
            ('社宅・寮費',       payslip.company_housing_fee),
            ('組合費',           payslip.union_fee),
            ('共済会費',         payslip.mutual_aid_fee),
            ('持株会拠出金',     payslip.employee_stock_contribution),
            ('その他控除',       payslip.other_deductions),
            ('所得税',           payslip.income_tax),
            ('住民税',           payslip.resident_tax),
        ]

        max_rows = max(len(PAYMENTS), len(DEDUCTIONS))
        rows = [['支給項目', '金額', '控除項目', '金額']]
        for i in range(max_rows):
            lp, vp = PAYMENTS[i] if i < len(PAYMENTS) else ('', '')
            ld, vd = DEDUCTIONS[i] if i < len(DEDUCTIONS) else ('', '')
            rows.append([lp, fmt(vp) if vp != '' else '', ld, fmt(vd) if vd != '' else ''])

        rows.append(['支給合計', fmt(payslip.gross_salary), '控除合計', fmt(payslip.total_deductions)])

        tbl = Table(rows, colWidths=[40*mm, 30*mm, 40*mm, 30*mm])
        ts  = [
            ('FONTNAME',   (0, 0), (-1, -1), FONT),
            ('FONTSIZE',   (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1,  0), BLUE_H),
            ('TEXTCOLOR',  (0, 0), (-1,  0), colors.white),
            ('GRID',       (0, 0), (-1, -1), 0.4, colors.grey),
            ('ALIGN',      (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN',      (3, 0), (3, -1), 'RIGHT'),
            ('ROWHEIGHT',  (0, 0), (-1, -1), 13),
            # 社会保険料合計行
            ('BACKGROUND', (2, 5), (3, 5), LBLUE),
            # 合計行
            ('BACKGROUND', (0, -1), (1, -1), GREEN),
            ('BACKGROUND', (2, -1), (3, -1), RED_),
            ('FONTNAME',   (0, -1), (-1, -1), FONT),
        ]
        tbl.setStyle(TableStyle(ts))
        elements += [tbl, Spacer(1, 6*mm)]

        # 差引支給額
        net_data = [[f'差引支給額（手取）: {fmt(payslip.net_salary)}']]
        net_tbl  = Table(net_data, colWidths=[140*mm])
        net_tbl.setStyle(TableStyle([
            ('FONTNAME',   (0, 0), (-1, -1), FONT),
            ('FONTSIZE',   (0, 0), (-1, -1), 13),
            ('BACKGROUND', (0, 0), (-1, -1), BLUE_H),
            ('TEXTCOLOR',  (0, 0), (-1, -1), colors.white),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('ROWHEIGHT',  (0, 0), (-1, -1), 20),
        ]))
        elements.append(net_tbl)

        doc.build(elements)
        buf.seek(0)

        filename = f'payslip_{payslip.year}{payslip.month:02d}_{emp.employee_number}.pdf'
        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


def _calculate_payslip(employee: Employee, year: int, month: int) -> Payslip:
    """給与計算ロジック（自動計算可能な項目を算出し、その他は0で初期化）"""
    from datetime import date
    from django.db.models import Q

    today = date(year, month, 1)

    # ── 基本給（等級マスタ） ──
    grade_obj = (
        SalaryGrade.objects.filter(grade=employee.grade, valid_from__lte=today, valid_to__isnull=True).first()
        or SalaryGrade.objects.filter(grade=employee.grade, valid_from__lte=today, valid_to__gte=today).first()
        or SalaryGrade.objects.filter(grade=employee.grade).order_by('-valid_from').first()
    )
    base_salary = grade_obj.base_salary if grade_obj else 0

    # ── 社員手当マスタから個別手当を取得 ──
    allowances = EmployeeAllowance.objects.filter(
        employee=employee, valid_from__lte=today,
    ).filter(
        Q(valid_to__isnull=True) | Q(valid_to__gte=today)
    ).select_related('allowance')

    # 手当タイプ別に集計
    allowance_map = {
        'housing':            0, 'secondment':        0,
        'technical':          0, 'commute':            0,
        'family':             0, 'certification':     0,
        'position':           0, 'special':           0,
        'perfect_attendance': 0, 'diligence':         0,
        'extra_overtime':     0,
    }
    for ea in allowances:
        at = ea.allowance.allowance_type
        if at in allowance_map:
            allowance_map[at] += (ea.amount if ea.amount is not None else ea.allowance.amount)

    # ── 残業手当（勤怠から自動計算） ──
    records = AttendanceRecord.objects.filter(
        employee=employee, date__year=year, date__month=month, clock_out__isnull=False,
    )
    total_overtime_min = sum(r.overtime_minutes for r in records)
    hourly_rate        = Decimal(str(base_salary)) / 160
    overtime_pay       = int(hourly_rate * Decimal(str(total_overtime_min / 60)) * OVERTIME_RATE)

    # ── 勤怠日数 ──
    work_days = records.filter(clock_in__isnull=False, clock_out__isnull=False).count()

    # ── 支給合計 ──
    total_allowances = sum(allowance_map.values())
    gross_salary     = base_salary + total_allowances + overtime_pay

    # ── 社会保険（概算） ──
    # 40歳以上は介護保険料あり
    from datetime import date as date_cls
    age = year - employee.birth_date.year - (
        (month, 1) < (employee.birth_date.month, employee.birth_date.day)
    )
    nursing_insurance    = int(Decimal(str(gross_salary)) * NURSING_INSURANCE_RATE) if age >= 40 else 0
    health_insurance     = int(Decimal(str(gross_salary)) * HEALTH_INSURANCE_RATE)
    pension              = int(Decimal(str(gross_salary)) * PENSION_RATE)
    employment_insurance = int(Decimal(str(gross_salary)) * EMPLOYMENT_INSURANCE_RATE)
    social_insurance_total = health_insurance + pension + employment_insurance + nursing_insurance

    # ── 所得税（簡易概算 5%） ──
    taxable    = gross_salary - social_insurance_total
    income_tax = int(Decimal(str(max(0, taxable))) * Decimal('0.05'))

    # ── 控除合計・差引支給額 ──
    total_deductions = social_insurance_total + income_tax  # 住民税等は人事が手動設定
    net_salary       = max(0, gross_salary - total_deductions)

    payslip, _ = Payslip.objects.update_or_create(
        employee=employee, year=year, month=month,
        defaults=dict(
            base_salary                  = base_salary,
            technical_allowance          = allowance_map['technical'],
            secondment_allowance         = allowance_map['secondment'],
            housing_allowance            = allowance_map['housing'],
            overtime_pay                 = overtime_pay,
            commute_allowance            = allowance_map['commute'],
            family_allowance             = allowance_map['family'],
            certification_allowance      = allowance_map['certification'],
            position_allowance           = allowance_map['position'],
            special_allowance            = allowance_map['special'],
            perfect_attendance_allowance = allowance_map['perfect_attendance'],
            diligence_allowance          = allowance_map['diligence'],
            extra_overtime_pay           = allowance_map['extra_overtime'],
            total_allowances             = total_allowances,
            gross_salary                 = gross_salary,
            health_insurance             = health_insurance,
            pension                      = pension,
            employment_insurance         = employment_insurance,
            nursing_insurance            = nursing_insurance,
            social_insurance_total       = social_insurance_total,
            income_tax                   = income_tax,
            total_deductions             = total_deductions,
            net_salary                   = net_salary,
            work_days                    = work_days,
            status                       = Payslip.Status.DRAFT,
        )
    )
    return payslip
