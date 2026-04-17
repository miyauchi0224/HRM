import csv
import io
import os
import calendar
from datetime import date, datetime, time
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from apps.accounts.permissions import IsNotCustomer
from rest_framework.response import Response
from django.http import HttpResponse
from django.utils import timezone
from django.shortcuts import get_object_or_404
from .models import AttendanceRecord, AttendanceProjectRecord, AttendanceModRequest, Project
from .serializers import (
    AttendanceRecordSerializer, ClockInSerializer, ClockOutSerializer,
    AttendanceModRequestSerializer, ProjectSerializer, AttendanceSummarySerializer
)
from apps.notifications.models import Notification


OVERTIME_WARNING_MINUTES = 40 * 60   # 40時間
OVERTIME_ALERT_MINUTES   = 80 * 60   # 80時間（36協定）


class ProjectViewSet(viewsets.ModelViewSet):
    """
    社員がプロジェクトを作成・編集・削除できる。
    GET /api/v1/attendance/projects/          - 一覧（アクティブのみ）
    POST /api/v1/attendance/projects/         - 作成
    PATCH /api/v1/attendance/projects/{id}/   - 更新
    DELETE /api/v1/attendance/projects/{id}/  - 削除（論理削除）
    """
    serializer_class   = ProjectSerializer
    permission_classes = [IsNotCustomer]

    def get_queryset(self):
        return Project.objects.filter(is_active=True)

    def perform_destroy(self, instance):
        # 物理削除ではなく論理削除（is_active=False）
        instance.is_active = False
        instance.save()


class AttendanceViewSet(viewsets.ModelViewSet):
    serializer_class   = AttendanceRecordSerializer
    permission_classes = [IsNotCustomer]
    http_method_names  = ['get', 'patch', 'head', 'options']  # PUT・DELETE は不要

    def get_queryset(self):
        user = self.request.user
        qs   = AttendanceRecord.objects.select_related('employee').prefetch_related(
            'project_records__project'
        )
        year_month = self.request.query_params.get('year_month')

        # 管理職・人事は部下/全員を参照可
        emp_id = self.request.query_params.get('employee_id')
        if user.is_manager and emp_id:
            qs = qs.filter(employee_id=emp_id)
        else:
            qs = qs.filter(employee__user=user)

        if year_month:
            year, month = year_month.split('-')
            qs = qs.filter(date__year=year, date__month=month)
        return qs.order_by('-date')

    def perform_update(self, serializer):
        """自分のレコードのみ編集可。保存時に status を MODIFIED にセット"""
        obj = self.get_object()
        if obj.employee.user != self.request.user and not self.request.user.is_manager:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('自身の打刻記録のみ編集できます')
        serializer.save(status=AttendanceRecord.Status.MODIFIED)

    @action(detail=False, methods=['post'], url_path='clock-in')
    def clock_in(self, request):
        """POST /api/v1/attendance/clock-in/"""
        serializer = ClockInSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        employee = getattr(request.user, 'employee', None)
        if not employee:
            return Response({'error': '社員情報が登録されていません。管理者にお問い合わせください。'},
                            status=status.HTTP_404_NOT_FOUND)
        today    = date.today()

        # 当日の記録が既にあればエラー
        if AttendanceRecord.objects.filter(employee=employee, date=today).exists():
            return Response({'error': '本日はすでに出勤打刻済みです'}, status=status.HTTP_409_CONFLICT)

        record = AttendanceRecord.objects.create(
            employee = employee,
            date     = today,
            clock_in = timezone.localtime().time(),
            note     = serializer.validated_data.get('note', ''),
        )
        return Response(AttendanceRecordSerializer(record).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='clock-out')
    def clock_out(self, request):
        """POST /api/v1/attendance/clock-out/"""
        serializer = ClockOutSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        employee = getattr(request.user, 'employee', None)
        if not employee:
            return Response({'error': '社員情報が登録されていません。管理者にお問い合わせください。'},
                            status=status.HTTP_404_NOT_FOUND)
        today    = date.today()
        record   = AttendanceRecord.objects.filter(
            employee=employee, date=today, clock_out__isnull=True
        ).first()

        if not record:
            return Response({'error': '出勤記録がありません'}, status=status.HTTP_404_NOT_FOUND)

        record.clock_out     = timezone.localtime().time()
        record.break_minutes = serializer.validated_data['break_minutes']
        record.status        = AttendanceRecord.Status.CONFIRMED
        record.save()

        # 残業アラートチェック
        self._check_overtime_alert(employee, record)

        return Response(AttendanceRecordSerializer(record).data)

    def _check_overtime_alert(self, employee, record):
        """月間残業時間が閾値を超えたら通知"""
        # list() で一度だけ評価してループ
        records = list(AttendanceRecord.objects.filter(
            employee=employee,
            date__year=record.date.year,
            date__month=record.date.month,
        ))
        total_overtime = sum(r.overtime_minutes for r in records)
        if total_overtime >= OVERTIME_ALERT_MINUTES:
            Notification.send(
                user=employee.user,
                type_=Notification.NotificationType.OVERTIME_ALERT,
                title='【緊急】残業時間アラート',
                message=f'今月の残業時間が{total_overtime // 60}時間を超えました。36協定の上限に達しています。',
                related_url='/attendance/report',
            )
        elif total_overtime >= OVERTIME_WARNING_MINUTES:
            Notification.send(
                user=employee.user,
                type_=Notification.NotificationType.OVERTIME_ALERT,
                title='残業時間警告',
                message=f'今月の残業時間が{total_overtime // 60}時間になりました。',
                related_url='/attendance/report',
            )

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """GET /api/v1/attendance/summary/?year_month=2026-04"""
        year_month = request.query_params.get('year_month',
                                               date.today().strftime('%Y-%m'))
        year, month = year_month.split('-')
        employee    = getattr(request.user, 'employee', None)
        if not employee:
            return Response({'error': '社員情報が登録されていません。管理者にお問い合わせください。'},
                            status=status.HTTP_404_NOT_FOUND)
        # list() で一度だけDBに問い合わせ（queryset の多重評価を防ぐ）
        records_list = list(AttendanceRecord.objects.filter(
            employee=employee, date__year=year, date__month=month,
            clock_in__isnull=False, clock_out__isnull=False,
        ))
        total_work     = sum(r.work_minutes for r in records_list)
        total_overtime = sum(r.overtime_minutes for r in records_list)
        return Response({
            'year_month':             year_month,
            'total_work_days':        len(records_list),
            'total_work_minutes':     total_work,
            'total_overtime_minutes': total_overtime,
            'overtime_alert':         total_overtime >= OVERTIME_ALERT_MINUTES,
        })

    @action(detail=False, methods=['get'], url_path='template')
    def template(self, request):
        """
        GET /api/v1/attendance/template/?year=2026[&employee_id=xxx]
        OC 作業表テンプレートに勤怠データを書き込んで XLSX でダウンロード
        """
        import openpyxl

        year   = int(request.query_params.get('year', date.today().year))
        emp_id = request.query_params.get('employee_id')
        if emp_id and request.user.is_manager:
            from apps.employees.models import Employee
            employee = Employee.objects.filter(id=emp_id).first()
        else:
            employee = getattr(request.user, 'employee', None)
        if not employee:
            return Response({'error': '社員情報が見つかりません'}, status=status.HTTP_404_NOT_FOUND)

        # テンプレートを読み込む（ボリュームマウントで /app/template/ に配置済み）
        template_path = os.path.join(os.path.dirname(__file__), '..', '..', 'template', 'OC作業表（名前）2026年度.xlsx')
        template_path = os.path.normpath(template_path)
        try:
            wb = openpyxl.load_workbook(template_path)
        except FileNotFoundError:
            return Response({'error': 'テンプレートが見つかりません: ' + template_path},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # 全勤怠データを日付→レコードのマップに変換（年をまたぐ可能性があるため全件取得）
        existing: dict[date, AttendanceRecord] = {
            r.date: r
            for r in AttendanceRecord.objects.filter(employee=employee)
        }

        # OC作業表 カラム位置（1始まり Excel 列番号）
        COL_DATE      = 2   # B: 日付
        COL_CLOCK_IN  = 22  # V: 始業
        COL_CLOCK_OUT = 23  # W: 終業
        COL_BREAK     = 24  # X: 休憩
        COL_NOTE      = 21  # U: 備考
        DATA_START    = 6   # 6行目からデータ
        SKIP_SHEETS   = {'件番'}

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue
            ws = wb[sheet_name]
            for row_num in range(DATA_START, ws.max_row + 1):
                cell_date = ws.cell(row=row_num, column=COL_DATE).value
                if cell_date is None:
                    continue
                # セル値が datetime または date のときのみ処理
                if isinstance(cell_date, datetime):
                    d = cell_date.date()
                elif isinstance(cell_date, date):
                    d = cell_date
                else:
                    continue

                rec = existing.get(d)
                if rec:
                    ws.cell(row=row_num, column=COL_CLOCK_IN).value  = rec.clock_in  or None
                    ws.cell(row=row_num, column=COL_CLOCK_OUT).value = rec.clock_out or None
                    if rec.break_minutes:
                        h, m = divmod(rec.break_minutes, 60)
                        ws.cell(row=row_num, column=COL_BREAK).value = time(h, m)
                    pj_codes = ' / '.join(pr.project.code for pr in rec.project_records.all())
                    ws.cell(row=row_num, column=COL_NOTE).value = pj_codes or rec.note or None
                else:
                    # データがない日はクリア
                    ws.cell(row=row_num, column=COL_CLOCK_IN).value  = None
                    ws.cell(row=row_num, column=COL_CLOCK_OUT).value = None
                    ws.cell(row=row_num, column=COL_BREAK).value     = None
                    ws.cell(row=row_num, column=COL_NOTE).value      = None

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'OC作業表（{employee.full_name}）{year}年度.xlsx'
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'], url_path='csv-export')
    def csv_export(self, request):
        """
        GET /api/v1/attendance/csv-export/?year_month=2026-04
        OC作業表と同じ列構造で CSV 出力（件番・始業・終業・休憩・就業時間・備考）
        """
        year_month = request.query_params.get('year_month', date.today().strftime('%Y-%m'))
        year, month = map(int, year_month.split('-'))
        employee = getattr(request.user, 'employee', None)
        if not employee:
            return Response({'error': '社員情報が登録されていません。'}, status=status.HTTP_404_NOT_FOUND)

        days_in_month = calendar.monthrange(year, month)[1]
        records_map   = {
            r.date: r
            for r in AttendanceRecord.objects.filter(
                employee=employee, date__year=year, date__month=month,
            )
        }
        weekdays = ['月', '火', '水', '木', '金', '土', '日']

        output = io.StringIO()
        writer = csv.writer(output)

        # ===== ヘッダー（OC作業表と同じ項目） =====
        writer.writerow(['氏名', employee.full_name])
        writer.writerow(['社員番号', employee.employee_number])
        writer.writerow(['部署', employee.department])
        writer.writerow([])
        writer.writerow(['日付', '曜日', '件番', '始業', '終業', '休憩', '就業時間(h)', '備考'])

        total_work = 0
        work_days  = 0
        for day in range(1, days_in_month + 1):
            d  = date(year, month, day)
            wd = weekdays[d.weekday()]
            r  = records_map.get(d)
            if r:
                ci   = r.clock_in.strftime('%H:%M')  if r.clock_in  else ''
                co   = r.clock_out.strftime('%H:%M') if r.clock_out else ''
                brk  = f'{r.break_minutes // 60}:{r.break_minutes % 60:02d}' if r.break_minutes else ''
                work = round(r.work_minutes / 60, 2) if r.work_minutes else ''
                pj   = ' / '.join(f'{pr.project.code}({pr.minutes}m)' for pr in r.project_records.all())
                note = r.note or ''
                if r.clock_in and r.clock_out:
                    total_work += r.work_minutes
                    work_days  += 1
            else:
                ci = co = brk = work = pj = note = ''
            writer.writerow([d.strftime('%Y/%m/%d'), wd, pj, ci, co, brk, work, note])

        writer.writerow([])
        writer.writerow(['出勤日数', f'{work_days}日',
                         '就業時間合計', f'{total_work // 60}:{total_work % 60:02d}'])

        content  = '\ufeff' + output.getvalue()  # UTF-8 BOM（Excel 文字化け防止）
        filename = f'OC作業表（{employee.full_name}）{year_month}.csv'
        response = HttpResponse(content, content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'], url_path='pdf-export')
    def pdf_export(self, request):
        """
        GET /api/v1/attendance/pdf-export/?year_month=2026-04
        OC作業表と同じ列構造で PDF 出力
        """
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from config.pdf_fonts import register_fonts
        register_fonts()  # 起動時登録済みだが冪等なので安全

        year_month = request.query_params.get('year_month', date.today().strftime('%Y-%m'))
        year, month = map(int, year_month.split('-'))
        employee = getattr(request.user, 'employee', None)
        if not employee:
            return Response({'error': '社員情報が登録されていません。'}, status=status.HTTP_404_NOT_FOUND)

        days_in_month = calendar.monthrange(year, month)[1]
        records_map   = {
            r.date: r
            for r in AttendanceRecord.objects.filter(
                employee=employee, date__year=year, date__month=month,
            )
        }
        weekdays = ['月', '火', '水', '木', '金', '土', '日']

        FONT = 'HeiseiKakuGo-W5'

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=12 * mm, rightMargin=12 * mm,
            topMargin=12 * mm, bottomMargin=12 * mm,
        )

        title_s = ParagraphStyle('t', fontName=FONT, fontSize=13, leading=18, spaceAfter=4)
        sub_s   = ParagraphStyle('s', fontName=FONT, fontSize=9,  leading=13)

        # 集計
        records_list  = list(records_map.values())
        total_work    = sum(r.work_minutes    for r in records_list)
        work_days     = sum(1 for r in records_list if r.clock_in and r.clock_out)

        elements = [
            Paragraph(f'OC作業表　{year}年{month}月', title_s),
            Paragraph(
                f'氏名: {employee.full_name}　社員番号: {employee.employee_number}　部署: {employee.department}',
                sub_s,
            ),
            Paragraph(
                f'出勤日数: {work_days}日　就業時間合計: {total_work // 60}:{total_work % 60:02d}',
                sub_s,
            ),
            Spacer(1, 5 * mm),
        ]

        # ===== テーブル（OC作業表と同じ列） =====
        header = ['日付', '曜', '件番', '始業', '終業', '休憩', '就業時間', '備考']
        rows   = [header]

        HEADER_BG  = colors.HexColor('#2E75B6')
        SAT_BG     = colors.HexColor('#D6E4FF')
        SUN_BG     = colors.HexColor('#FFE4E4')

        for day in range(1, days_in_month + 1):
            d  = date(year, month, day)
            wd = weekdays[d.weekday()]
            r  = records_map.get(d)
            if r:
                ci   = r.clock_in.strftime('%H:%M')  if r.clock_in  else ''
                co   = r.clock_out.strftime('%H:%M') if r.clock_out else ''
                brk  = f'{r.break_minutes // 60}:{r.break_minutes % 60:02d}' if r.break_minutes else ''
                work = f'{r.work_minutes // 60}:{r.work_minutes % 60:02d}' if r.work_minutes else ''
                pj   = ' / '.join(f'{pr.project.code}({pr.minutes}m)' for pr in r.project_records.all())
                note = r.note or ''
            else:
                ci = co = brk = work = pj = note = ''
            rows.append([d.strftime('%m/%d'), wd, pj, ci, co, brk, work, note])

        # 月次合計行
        rows.append([
            '合計', '', '', '', '', '',
            f'{total_work // 60}:{total_work % 60:02d}',
            f'{work_days}日出勤',
        ])

        col_widths = [16*mm, 10*mm, 28*mm, 16*mm, 16*mm, 14*mm, 16*mm, 60*mm]
        tbl = Table(rows, colWidths=col_widths, repeatRows=1)

        ts = [
            ('FONTNAME',   (0, 0), (-1, -1), FONT),
            ('FONTSIZE',   (0, 0), (-1,  0), 8),
            ('FONTSIZE',   (0, 1), (-1, -1), 7),
            ('TEXTCOLOR',  (0, 0), (-1,  0), colors.white),
            ('BACKGROUND', (0, 0), (-1,  0), HEADER_BG),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN',      (7, 1), (7,  -1), 'LEFT'),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#AAAAAA')),
            ('ROWHEIGHT',  (0, 0), (-1, -1), 6.5 * mm),
            # 合計行
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E2EFDA')),
            ('FONTSIZE',   (0, -1), (-1, -1), 8),
            ('FONTNAME',   (0, -1), (-1, -1), FONT),
        ]
        for i, day in enumerate(range(1, days_in_month + 1), start=1):
            d = date(year, month, day)
            if d.weekday() == 5:
                ts.append(('BACKGROUND', (0, i), (-1, i), SAT_BG))
            elif d.weekday() == 6:
                ts.append(('BACKGROUND', (0, i), (-1, i), SUN_BG))

        tbl.setStyle(TableStyle(ts))
        elements.append(tbl)

        doc.build(elements)
        buf.seek(0)

        filename = f'OC作業表（{employee.full_name}）{year}-{month:02d}.pdf'
        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['post'], url_path='upload',
            parser_classes=[MultiPartParser, FormParser])
    def upload(self, request):
        """
        POST /api/v1/attendance/upload/
        XLSX または CSV をアップロードして勤怠データを一括登録・更新
        CSV 列順: 日付, 曜日, 出勤時刻, 退勤時刻, 休憩(分), 労働時間(分), 残業時間(分), プロジェクト, 備考
        """
        import openpyxl

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'ファイルが指定されていません'}, status=status.HTTP_400_BAD_REQUEST)

        emp_id = request.data.get('employee_id')
        if emp_id and request.user.is_manager:
            from apps.employees.models import Employee
            employee = Employee.objects.filter(id=emp_id).first()
        else:
            employee = getattr(request.user, 'employee', None)

        if not employee:
            return Response({'error': '社員情報が見つかりません'}, status=status.HTTP_404_NOT_FOUND)

        # 時刻パース共通関数
        def parse_time(val):
            if not val:
                return None
            if isinstance(val, time):
                return val
            if isinstance(val, datetime):
                return val.time()
            s = str(val).strip()
            if ':' in s:
                parts = s.split(':')
                return time(int(parts[0]), int(parts[1]))
            return None

        # 1行分のデータを DB に登録/更新する共通関数
        def upsert_row(date_val, clock_in_str, clock_out_str, break_min, project_code, note, context=''):
            if isinstance(date_val, datetime):
                d = date_val.date()
            elif isinstance(date_val, date):
                d = date_val
            elif isinstance(date_val, (int, float)):
                # Excel シリアル日付番号 → date に変換
                # openpyxl の data_only=True でフォーミュラキャッシュが整数のとき
                try:
                    from openpyxl.utils.datetime import from_excel
                    converted = from_excel(int(date_val))
                    d = converted.date() if hasattr(converted, 'date') else converted
                    # 1900年代など明らかに不正な場合はスキップ
                    if d.year < 2000:
                        return None
                except Exception:
                    raise ValueError(f'日付形式が不正: {date_val}')
            else:
                date_str = str(date_val).strip()
                for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%m/%d/%Y', '%m/%d'):
                    try:
                        d = datetime.strptime(date_str, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    raise ValueError(f'日付形式が不正: {date_val}')

            ci = parse_time(clock_in_str)
            co = parse_time(clock_out_str)
            bm = int(float(str(break_min))) if break_min not in (None, '', '-') else 60

            project = None
            if project_code and str(project_code).strip() not in ('', '-'):
                project = Project.objects.filter(code=str(project_code).strip()).first()

            record, is_created = AttendanceRecord.objects.update_or_create(
                employee=employee,
                date=d,
                defaults=dict(
                    clock_in=ci,
                    clock_out=co,
                    break_minutes=bm,
                    note=str(note) if note else '',
                    status=AttendanceRecord.Status.CONFIRMED if ci and co else AttendanceRecord.Status.DRAFT,
                )
            )
            # プロジェクトコードがあれば作業時間を全労働時間として登録
            if project and ci and co:
                from datetime import datetime as _dt, date as _d
                _ci = _dt.combine(_d.today(), ci)
                _co = _dt.combine(_d.today(), co)
                work = max(0, int((_co - _ci).total_seconds() / 60) - bm)
                AttendanceProjectRecord.objects.update_or_create(
                    attendance=record,
                    project=project,
                    defaults={'minutes': work},
                )
            return is_created

        created = 0
        updated = 0
        errors  = []

        filename = file.name.lower()

        # ===== CSV =====
        if filename.endswith('.csv'):
            try:
                content = file.read().decode('utf-8-sig')  # BOM 付きも対応
                reader  = csv.reader(io.StringIO(content))
                for line_no, row in enumerate(reader, start=1):
                    if line_no == 1:
                        continue  # ヘッダー行をスキップ
                    if not row or not row[0].strip():
                        continue
                    try:
                        # 列: 日付, 曜日, 出勤, 退勤, 休憩(分), 労働(分), 残業(分), PJ, 備考
                        date_val     = row[0]
                        clock_in_str = row[2] if len(row) > 2 else ''
                        clock_out_str= row[3] if len(row) > 3 else ''
                        break_min    = row[4] if len(row) > 4 else 60
                        project_code = row[7] if len(row) > 7 else ''
                        note         = row[8] if len(row) > 8 else ''
                        result = upsert_row(date_val, clock_in_str, clock_out_str,
                                           break_min, project_code, note,
                                           context=f'行{line_no}')
                        if result is None:
                            pass  # 変換不可な日付（1900年代等）はスキップ
                        elif result:
                            created += 1
                        else:
                            updated += 1
                    except Exception as e:
                        errors.append(f'CSV {line_no}行目: {str(e)}')
            except UnicodeDecodeError:
                return Response({'error': 'CSV のエンコードは UTF-8 または Shift-JIS にしてください'}, status=status.HTTP_400_BAD_REQUEST)

        # ===== XLSX (OC作業表テンプレート対応) =====
        else:
            try:
                wb = openpyxl.load_workbook(file, data_only=True)
            except Exception:
                return Response({'error': '無効なXLSXファイルです'}, status=status.HTTP_400_BAD_REQUEST)

            def safe(row, idx, default=None):
                """tuple/list の安全なインデックスアクセス"""
                try:
                    return row[idx]
                except (IndexError, TypeError):
                    return default

            def parse_break_minutes(val):
                """
                休憩欄の値を分に変換
                '01:15:00' → 75, datetime.time(1,15) → 75, 1.25 → 75 など
                """
                if val is None or str(val).strip() in ('', '0', '0.0'):
                    return 60  # デフォルト60分
                if isinstance(val, (time, datetime)):
                    t = val.time() if isinstance(val, datetime) else val
                    return t.hour * 60 + t.minute
                s = str(val).strip()
                if ':' in s:
                    parts = s.split(':')
                    try:
                        return int(parts[0]) * 60 + int(parts[1])
                    except (ValueError, IndexError):
                        pass
                try:
                    # 時間数（小数）として解釈 e.g. 1.25 → 75分
                    return round(float(s) * 60)
                except ValueError:
                    return 60

            # OC作業表のカラム定義（0始まりインデックス）
            # 件番シートはプロジェクト一覧なので読み飛ばす
            OC_SKIP_SHEETS   = {'件番'}
            OC_DATE_COL      = 1   # B列: 日付
            OC_CLOCK_IN_COL  = 21  # V列: 始業
            OC_CLOCK_OUT_COL = 22  # W列: 終業
            OC_BREAK_COL     = 23  # X列: 休憩
            OC_NOTE_COL      = 20  # U列: 備考
            OC_PJCODE_ROW    = 3   # 件番ヘッダー行（0始まり → 実際の行3 = インデックス2）
            OC_DATA_START    = 6   # データ開始行（OC作業表は6行目から）

            for sheet in wb.worksheets:
                # 件番シート（プロジェクト一覧）はスキップ
                if sheet.title in OC_SKIP_SHEETS:
                    continue

                # プロジェクトコードを件番ヘッダー行から取得（行3=インデックス2）
                pj_header = list(sheet.iter_rows(
                    min_row=OC_PJCODE_ROW, max_row=OC_PJCODE_ROW, values_only=True
                ))
                project_code = ''
                if pj_header:
                    # 件番が入っているセルを探す（B列以降の最初の非Noneかつ文字列）
                    for cell_val in (pj_header[0] or []):
                        if cell_val and isinstance(cell_val, str) and len(cell_val) >= 4:
                            project_code = str(cell_val).strip()
                            break

                # データ行を処理
                for row in sheet.iter_rows(min_row=OC_DATA_START, values_only=True):
                    try:
                        date_val = safe(row, OC_DATE_COL)
                        if not date_val:
                            continue
                        # 曜日セルがNoneや空なら集計行なのでスキップ
                        weekday = safe(row, 2)
                        if weekday is None:
                            continue

                        clock_in_str  = safe(row, OC_CLOCK_IN_COL)
                        clock_out_str = safe(row, OC_CLOCK_OUT_COL)
                        break_min     = parse_break_minutes(safe(row, OC_BREAK_COL))
                        note          = str(safe(row, OC_NOTE_COL, '') or '')

                        # 出勤時刻も退勤時刻もなければスキップ（休日・空白行）
                        if not clock_in_str and not clock_out_str:
                            continue

                        result = upsert_row(
                            date_val, clock_in_str, clock_out_str,
                            break_min, project_code, note,
                            context=f'シート「{sheet.title}」',
                        )
                        if result is None:
                            pass  # 変換不可な日付はスキップ
                        elif result:
                            created += 1
                        else:
                            updated += 1
                    except Exception as e:
                        errors.append(f'シート「{sheet.title}」: {str(e)}')

        return Response({'created': created, 'updated': updated, 'errors': errors})


class AttendanceModRequestViewSet(viewsets.ModelViewSet):
    serializer_class   = AttendanceModRequestSerializer
    permission_classes = [IsNotCustomer]

    def get_queryset(self):
        user = self.request.user
        if user.is_manager:
            return AttendanceModRequest.objects.all()
        return AttendanceModRequest.objects.filter(applicant__user=user)

    def perform_create(self, serializer):
        serializer.save(applicant=self.request.user.employee)
        employee = self.request.user.employee
        for manager in employee.managers.all():
            Notification.send(
                user=manager.user,
                type_=Notification.NotificationType.ATTENDANCE_MOD,
                title='打刻修正申請',
                message=f'{employee.full_name}さんから打刻修正申請が届きました。',
                related_url='/attendance/modification',
            )

    @action(detail=True, methods=['patch'], url_path='approve')
    def approve(self, request, pk=None):
        """PATCH /api/v1/attendance/modification-requests/{id}/approve/"""
        if not request.user.is_manager:
            return Response({'error': '権限がありません'}, status=status.HTTP_403_FORBIDDEN)
        mod = self.get_object()
        if mod.status != AttendanceModRequest.Status.PENDING:
            return Response({'error': '申請中のものだけ承認できます'}, status=status.HTTP_400_BAD_REQUEST)

        action_type = request.data.get('action', 'approve')
        if action_type == 'approve':
            mod.status      = AttendanceModRequest.Status.APPROVED
            mod.approver    = request.user.employee
            mod.approved_at = timezone.now()
            mod.save()
            # 勤怠記録に反映
            mod.attendance.clock_in  = mod.requested_clock_in
            mod.attendance.clock_out = mod.requested_clock_out
            mod.attendance.status    = AttendanceRecord.Status.MODIFIED
            mod.attendance.save()
            msg = '打刻修正申請が承認されました'
        else:
            mod.status = AttendanceModRequest.Status.REJECTED
            mod.save()
            msg = '打刻修正申請が却下されました'

        Notification.send(
            user=mod.applicant.user,
            type_=Notification.NotificationType.ATTENDANCE_MOD,
            title=msg,
            message=msg,
            related_url='/attendance/calendar',
        )
        return Response(AttendanceModRequestSerializer(mod).data)
